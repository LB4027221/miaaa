#!/usr/bin/python
# -*- coding: UTF-8 -*-

import sys
import json
import pandas as pd
import xlsxwriter
import os
import pymysql
import re

from datetime import datetime, timedelta, date
from xlsxwriter.utility import xl_rowcol_to_cell

from openpyxl.formula import Tokenizer
from openpyxl import load_workbook

msg = json.loads(sys.argv[1])
filename = msg['filename']
worksheetName = msg['ws']
data = msg['data']

xls = load_workbook(filename)

sheet = xls[worksheetName]

for rowIdx, row in enumerate(data):
  for colIdx, column in enumerate(row):
    sheet.cell(row=rowIdx + 1, column=colIdx + 1).value = column

xls.save(filename)

msg = { 'success': True }

sys.stdout.write(json.dumps(msg))
sys.exit()
