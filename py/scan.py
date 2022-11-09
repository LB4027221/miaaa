#!/usr/bin/python
# -*- coding: UTF-8 -*-

import sys
import json
import pandas as pd
import xlsxwriter
import os
import pymysql
from openpyxl import load_workbook

msg = json.loads(sys.argv[1])
filename = msg['filename']
worksheetName = msg['worksheetName']
sql = msg['sql']
output = msg['output']

xls = load_workbook(filename)

sheets = {}

writer = pd.ExcelWriter(output, engine='xlsxwriter')
workbook = writer.book

conn = pymysql.connect(host='rm-bp1enc0yrzt3zu60do.mysql.rds.aliyuncs.com', port=3306, user='sxc_test', passwd='Songxiaocai2015')

for sheet_name in xls.sheetnames:
  sheets[sheet_name] = xls[sheet_name]

sheets[worksheetName] = pd.read_sql(sql, conn)

for sheet_name in xls.sheetnames:
  sheets[sheet_name] = pd.DataFrame(sheets[sheet_name].values)
  sheets[sheet_name].to_excel(writer, index=False, sheet_name=sheet_name, header=0)

writer.save()

sys.stdout.write('写入完成')
sys.exit()
