#!/usr/bin/python
# -*- coding: UTF-8 -*-

import sys
import json
import pandas as pd
import xlsxwriter
import os
import pymysql
import re

from datetime import datetime, timedelta, date
from xlsxwriter.utility import xl_rowcol_to_cell

from openpyxl.formula import Tokenizer
from openpyxl import load_workbook

today = datetime.now()

def validate(date_text):
  if (type(date_text) is datetime):
    return date_text.strftime('%Y/%m/%d')

  return date_text

def writeData(rowIdx, cellIdx, value, ws, date_format):
  notation = xl_rowcol_to_cell(rowIdx, cellIdx)
  if (type(value) == str and len(value) > 1 and value[0] == '='):
    ws.write_formula(notation, value)
  else:
    ws.write(notation, validate(value))

def dashreplByNum(num):
  def dashrepl(matchobj):
    matched = matchobj.group(0)
    return matched.replace("2", num)

  return dashrepl

def writeDataByTpl(rowIdx, cellIdx, ws, tpl, date_format):
  notation = xl_rowcol_to_cell(rowIdx, cellIdx)
  notation_num = re.findall(r'\d+', notation)[0]
  value = tpl[cellIdx]

  if (value == None):
    return

  if (type(value) == str and len(value) > 1 and value[0] == '='):
    reg = '[A-Z]2'
    dashreplFn = dashreplByNum(notation_num)
    new_formula = re.sub(reg, dashreplFn, value)
    ws.write_formula(rowIdx, cellIdx, new_formula)

  else:
    ws.write(rowIdx, cellIdx, validate(value))

def getFirstRowFormula(ws):
  col = []
  firstRow = ws.iter_rows(min_row=2, max_row=2)
  for row in firstRow:
    for cell in row:
      col.append(cell.value)
  return col

def only_copy(source, ws, date_format):
  for rowIdx, row in enumerate(source.iter_rows()):
    for cellIdx, cell in enumerate(row):
      writeData(rowIdx, cellIdx, cell.value, ws, date_format)

def copy_col(source, ws, date_format):
  for rowIdx, row in enumerate(source.iter_rows(max_row=1)):
    for cellIdx, cell in enumerate(row):
      writeData(rowIdx, cellIdx, cell.value, ws, date_format)

def fill_ws(source, ws, dataLen, date_format):
  tpl = getFirstRowFormula(source)

  for rowIdx in range(dataLen):
    for cellIdx in range(len(tpl)):
      writeDataByTpl(rowIdx + 1, cellIdx, ws, tpl, date_format)

def get_date(max, min = 0):
  days = []
  for i in range(min, max): 
    date_N_days_ago = datetime.now() - timedelta(days=i)
    days.append(date_N_days_ago.strftime('%Y/%m/%d'))

  days.reverse()

  return days

def last_day_of_month(any_day):
  next_month = any_day.replace(day=28) + timedelta(days=4)
  return next_month - timedelta(days=next_month.day)

def get_last_month():
  days = []
  todayDate = date.today()
  first_day = todayDate.replace(months=-1,day=1)
  last_day = last_day_of_month(todayDate)

  step = timedelta(days=1)

  while first_day <= last_day:
    days.append(first_day)
    first_day += step

  days.reverse()

  return days

def get_last_week():
  days = []
  todayDate = date.today()
  first_day = todayDate.replace(weeks=-1,day=1)
  last_day = last_day_of_month(todayDate)

  step = timedelta(days=1)

  while first_day <= last_day:
    days.append(first_day)
    first_day += step

  days.reverse()

  return days

def get_past_week():
  return get_date(7)

def get_this_week():
  maxRange = today.weekday()
  return get_date(maxRange)

def get_this_month():
  maxRange = today.day
  return get_date(maxRange)

def get_past_month():
  return get_date(30)

def get_today():
  return get_date(1)

def write_col_by_date(source, ws, date_format, days):
  for rowIdx, row in enumerate(days):
    writeData(rowIdx + 1, 0, row, ws, date_format)

def write_by_date(source, ws, date_format, days):
  copy_col(source, ws, date_format)
  fill_ws(source, ws, len(days), date_format)
  write_col_by_date(source, ws, date_format, days)

def get_days_by_range(range):
  data = pd.date_range(range[0], range[1])

  return data

def filterParser(conf, source):
  res = source.copy()

  for item in conf:
    for condition in item['conditions']:
      if condition["name"] == "by_value":
        res = res[res[item["columnName"]].isin(condition["args"][0])]

  return res

msg = json.loads(sys.argv[1])
filename = msg['filename']
worksheetName = msg['worksheetName']
sql = msg['sql']
output = msg['output']
host = msg['host']
user = msg['user']
passwd = msg['passwd']
conf = msg['conf']
clearSource = msg['clearSource']

xls = load_workbook(filename)

sheets = {}

writer = pd.ExcelWriter(output, engine='xlsxwriter')
workbook = writer.book
date_format = workbook.add_format({'num_format': 'yyyy/mm/dd'})

conn = pymysql.connect(host=host, port=3306, user=user, passwd=passwd)
sheets[worksheetName] = pd.read_sql(sql, conn)

main_excel = sheets[worksheetName]

if clearSource:
  main_excel = main_excel.head(10)

main_excel.to_excel(writer, index=False, sheet_name=worksheetName)

source = sheets[worksheetName]

# describeSheet = sheets[worksheetName].describe()
# describeSheet.to_excel(writer, index=True, sheet_name='自动分析簿')

for sheet_name in xls.sheetnames:
  if sheet_name != worksheetName:
    if sheet_name.find('_去重') == -1 and sheet_name not in conf:
      sheets[sheet_name] = workbook.add_worksheet(sheet_name)

    if sheet_name.find('_填充') > -1:
      dataLen = len(sheets[worksheetName])
      copy_col(xls[sheet_name], sheets[sheet_name], date_format)
      fill_ws(xls[sheet_name], sheets[sheet_name], dataLen, date_format)

    if sheet_name in conf:
      sheet_name = sheet_name.strip()
      if 'timeHandler' in conf[sheet_name] and conf[sheet_name]['timeHandler'] == 'insertTime':
        days = get_days_by_range(conf[sheet_name]['dateRange'])
        write_by_date(xls[sheet_name], sheets[sheet_name], date_format, days)

      else:
        new_source = source.copy()

        if (len(new_source)):
          if 'timeHandler' in conf[sheet_name] and conf[sheet_name]['timeHandler'] == 'fitlerByTime':
            index = conf[sheet_name]['index']
            start = conf[sheet_name]['dateRange'][0]
            end = conf[sheet_name]['dateRange'][1]
            new_source[index] = pd.to_datetime(new_source[index])
            mask = (new_source[index] >= start) & (new_source[index] <= end)
            new_source = new_source.loc[mask]
            new_source[index] = new_source[index].dt.strftime("%Y/%m/%d")

          if 'filter' in conf[sheet_name]:
            new_source = filterParser(conf[sheet_name]['filter'], new_source)

          if 'unique' in conf[sheet_name]:
            new_source = new_source.drop_duplicates(subset=conf[sheet_name]['unique'])
  
        new_source.to_excel(writer, index=False, sheet_name=sheet_name)

    else:
      only_copy(xls[sheet_name], sheets[sheet_name], date_format)

writer.save()
workbook.close()

msg = { 'success': True }

sys.stdout.write(json.dumps(msg))
sys.exit()