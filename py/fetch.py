#!/usr/bin/python
# -*- coding: UTF-8 -*-

import sys
import json
import pandas as pd
import xlsxwriter
import os
import pymysql
import re

from datetime import datetime, timedelta
from xlsxwriter.utility import xl_rowcol_to_cell

from openpyxl.formula import Tokenizer
from openpyxl import load_workbook

today = datetime.now()

def validate(date_text):
  if (type(date_text) is datetime):
    return date_text.strftime('%Y/%m/%d')

  return date_text

def writeData(rowIdx, cellIdx, value, ws, date_format):
  notation = xl_rowcol_to_cell(rowIdx, cellIdx)
  if (type(value) == str and len(value) > 1 and value[0] == '='):
    ws.write_formula(notation, value)
  else:
    ws.write(notation, validate(value))

def dashreplByNum(num):
  def dashrepl(matchobj):
    matched = matchobj.group(0)
    return matched.replace("2", num)

  return dashrepl

def writeDataByTpl(rowIdx, cellIdx, ws, tpl, date_format):
  notation = xl_rowcol_to_cell(rowIdx, cellIdx)
  notation_num = re.findall(r'\d+', notation)[0]
  value = tpl[cellIdx]

  if (value == None):
    return

  if (type(value) == str and len(value) > 1 and value[0] == '='):
    reg = '[A-Z]2'
    dashreplFn = dashreplByNum(notation_num)
    new_formula = re.sub(reg, dashreplFn, value)
    ws.write_formula(rowIdx, cellIdx, new_formula)

  else:
    ws.write(rowIdx, cellIdx, validate(value))

def getFirstRowFormula(ws):
  col = []
  firstRow = ws.iter_rows(min_row=2, max_row=2)
  for row in firstRow:
    for cell in row:
      col.append(cell.value)
  return col

def only_copy(source, ws, date_format):
  for rowIdx, row in enumerate(source.iter_rows()):
    for cellIdx, cell in enumerate(row):
      writeData(rowIdx, cellIdx, cell.value, ws, date_format)

def copy_col(source, ws, date_format):
  for rowIdx, row in enumerate(source.iter_rows(max_row=1)):
    for cellIdx, cell in enumerate(row):
      writeData(rowIdx, cellIdx, cell.value, ws, date_format)

def fill_ws(source, ws, dataLen, date_format):
  tpl = getFirstRowFormula(source)

  for rowIdx in range(dataLen):
    for cellIdx in range(len(tpl)):
      writeDataByTpl(rowIdx + 1, cellIdx, ws, tpl, date_format)

def get_date(max, min = 0):
  days = []
  for i in range(min, max): 
    date_N_days_ago = datetime.now() - timedelta(days=i)
    days.append(date_N_days_ago.strftime('%Y/%m/%d'))

  return days

def get_last_week():
  return get_date(7)

def get_this_week():
  maxRange = today.weekday() + 1
  return get_date(maxRange)

def get_this_month():
  maxRange = today.day + 1
  return get_date(maxRange)

def get_last_month():
  return get_date(30)

def get_today():
  return get_date(1)

def write_col_by_date(source, ws, date_format, days):
  for rowIdx, row in enumerate(days):
    writeData(rowIdx + 1, 0, row, ws, date_format)

def write_by_date(source, ws, date_format, days):
  copy_col(source, ws, date_format)
  fill_ws(source, ws, len(days), date_format)
  write_col_by_date(source, ws, date_format, days)

msg = json.loads(sys.argv[1])
worksheetName = msg['worksheetName']
sql = msg['sql']
output = msg['output']
host = msg['host']
user = msg['user']
passwd = msg['passwd']

sheets = {}

writer = pd.ExcelWriter(output, engine='xlsxwriter')
workbook = writer.book

conn = pymysql.connect(host=host, port=3306, user=user, passwd=passwd)
sheets[worksheetName] = pd.read_sql(sql, conn)
sheets[worksheetName].to_excel(writer, index=False, sheet_name=worksheetName)

writer.save()
workbook.close()

msg = { 'success': True }

sys.stdout.write(json.dumps(msg))
sys.exit()
